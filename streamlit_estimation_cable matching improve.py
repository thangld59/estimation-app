# streamlit_estimation_app_final_quotation.py
# BuildWise - Estimation & Quotation (Excel only, with template)
# - Login / user management
# - Per-user price lists
# - Customers & company profile
# - Cable matching (same logic as last good version)
# - Quotation generation using quotation_template.xlsx
# - Quotation preview + Download + Save to history
# - No PDF generation

import streamlit as st
import pandas as pd
import os
import re
import json
from io import BytesIO
from datetime import datetime
from rapidfuzz import fuzz
from openpyxl import load_workbook

# ------------------------------
# Constants / paths
# ------------------------------
USERS_FILE = "users.json"
FORM_FOLDER = "shared_forms"
TEMPLATE_FILE = "quotation_template.xlsx"

DEFAULT_USERS = {
    "Admin123": {"password": "BuildWise2025", "role": "admin"},
    "User123": {"password": "User2025", "role": "user"}
}

# ------------------------------
# Users
# ------------------------------
def save_users(users):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, indent=2, ensure_ascii=False)

def load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return DEFAULT_USERS.copy()
    else:
        save_users(DEFAULT_USERS)
        return DEFAULT_USERS.copy()

USERS = load_users()

# ------------------------------
# Matching utilities (unchanged logic)
# ------------------------------
MAIN_SIZE_RE = re.compile(r"\b(\d{1,2})\s*[cC]?\s*[x×]\s*(\d{1,3}(?:\.\d+)?)\b")
AUX_RE = re.compile(
    r"\+\s*(?:([1-9]\d*)\s*[cC]?\s*[x×]\s*)?((?:pe|e|n))?(\d{1,3}(?:\.\d+)?)",
    flags=re.IGNORECASE,
)
MATERIAL_TOKEN_RE = re.compile(
    r"(cu|aluminium|al|xlpe|pvc|pe|lszh|hdpe|dsta|sta|swa)",
    flags=re.IGNORECASE,
)

def clean(text: str) -> str:
    text = str(text).lower()
    text = re.sub(r"0[,.]?6kv|1[,.]?0kv", "", text)
    text = text.replace("mm2", "").replace("mm²", "")
    text = text.replace("(", "").replace(")", "")
    text = text.replace("/", " ").replace(",", "")
    text = text.replace("-", " ")
    text = text.replace("cáp", "").replace("cable", "").replace("dây", "")
    text = re.sub(r"\s+", " ", text).strip()
    return text

def parse_cable_spec(text: str) -> dict:
    text = str(text).lower().replace("mm2", "").replace("mm²", "")
    text = re.sub(r"\s+", " ", text)

    main_match = MAIN_SIZE_RE.search(text)
    main_cores, main_size = None, None
    if main_match:
        try:
            main_cores = int(main_match.group(1))
        except Exception:
            main_cores = None
        try:
            main_size = float(main_match.group(2))
        except Exception:
            main_size = None

    aux_match = AUX_RE.search(text)
    aux_type = ""
    aux_cores = None
    aux_size = None
    if aux_match:
        cores_str = aux_match.group(1)
        type_str = aux_match.group(2)
        size_str = aux_match.group(3)

        if cores_str:
            try:
                aux_cores = int(cores_str)
            except Exception:
                aux_cores = None

        if type_str:
            t = type_str.strip().upper()
            if t in ["E", "PE"]:
                aux_type = "E"
            elif t == "N":
                aux_type = "N"

        try:
            aux_size = float(size_str)
        except Exception:
            aux_size = None

    main_key = (
        f"{int(main_cores)}x{int(main_size) if main_size and float(main_size).is_integer() else main_size}"
        if main_cores and main_size
        else ""
    )

    if aux_type and aux_size:
        aux_key = f"{aux_type}{int(aux_size) if aux_size and float(aux_size).is_integer() else aux_size}"
    elif aux_cores and aux_size:
        aux_key = f"{aux_cores}x{int(aux_size) if aux_size and float(aux_size).is_integer() else aux_size}"
    else:
        aux_key = ""

    full_key = f"{main_key}+{aux_key}" if main_key and aux_key else main_key
    return {
        "main_cores": main_cores,
        "main_size": main_size,
        "aux_type": aux_type,
        "aux_cores": aux_cores,
        "aux_size": aux_size,
        "main_key": main_key,
        "aux_key": aux_key,
        "full_key": full_key,
    }

def extract_material_structure_tokens(text: str):
    text = str(text).lower()
    tokens = MATERIAL_TOKEN_RE.findall(text)
    norm = []
    for t in tokens:
        tt = t.lower()
        if tt == "aluminium":
            norm.append("al")
        else:
            norm.append(tt)
    return norm

def material_structure_score(query_tokens, target_tokens):
    if not query_tokens and not target_tokens:
        return 100.0
    if not query_tokens or not target_tokens:
        return 0.0

    weights_map = {
        "cu": 1.0,
        "al": 1.0,
        "xlpe": 0.9,
        "pvc": 0.7,
        "lszh": 0.6,
        "pe": 0.5,
        "hdpe": 0.5,
        "dsta": 0.4,
        "sta": 0.4,
        "swa": 0.4,
    }

    q_set = list(dict.fromkeys(query_tokens))
    t_set = list(dict.fromkeys(target_tokens))

    match_score = 0.0
    possible_score = 0.0

    all_keys = list(dict.fromkeys(q_set + t_set))
    for k in all_keys:
        w = weights_map.get(k, 0.3)
        possible_score += w
        if k in q_set and k in t_set:
            match_score += w

    base = (match_score / possible_score) * 100.0 if possible_score > 0 else 0.0
    extra_in_target = len([k for k in t_set if k not in q_set])
    extra_in_query = len([k for k in q_set if k not in t_set])
    penalty = extra_in_target * 5.0 + extra_in_query * 2.0
    score = max(0.0, base - penalty)
    return score

def combined_match_score(
    query,
    q_main_key,
    q_aux_key,
    q_mats,
    row_combined,
    r_main_key,
    r_aux_key,
    r_mats,
    threshold,
    weights,
):
    size_score = 0.0
    cores_score = 0.0
    mat_score = 0.0

    if q_main_key and r_main_key:
        if q_main_key == r_main_key:
            size_score = 100.0
        else:
            size_score = fuzz.token_set_ratio(q_main_key, r_main_key)
    else:
        size_score = fuzz.partial_ratio(query, row_combined)

    if q_aux_key and r_aux_key:
        if q_aux_key == r_aux_key:
            cores_score = 100.0
        else:
            cores_score = fuzz.token_set_ratio(str(q_aux_key), str(r_aux_key))
    else:
        cores_score = 100.0 if (not q_aux_key and not r_aux_key) else 0.0

    mat_score = material_structure_score(q_mats, r_mats)

    final = (
        weights["size"] * size_score
        + weights["cores"] * cores_score
        + weights["material"] * mat_score
    )
    return final

# ------------------------------
# Streamlit setup & login
# ------------------------------
st.set_page_config(page_title="BuildWise", page_icon="📐", layout="wide")

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
    st.session_state["username"] = ""
    st.session_state["role"] = ""

def do_login(user: str, pwd: str):
    user = user.strip()
    if user in USERS and USERS[user]["password"] == pwd:
        st.session_state["logged_in"] = True
        st.session_state["username"] = user
        st.session_state["role"] = USERS[user].get("role", "user")
        return True
    return False

def do_logout():
    st.session_state["logged_in"] = False
    st.session_state["username"] = ""
    st.session_state["role"] = ""

if not st.session_state["logged_in"]:
    st.title("📐 BuildWise - Sign in")
    with st.form("login_form", clear_on_submit=False):
        u = st.text_input("Username")
        p = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Log in")
        if submitted:
            ok = do_login(u, p)
            if ok:
                st.success(
                    f"Logged in as {st.session_state['username']} ({st.session_state['role']})"
                )
                st.experimental_rerun()
            else:
                st.error("Invalid username or password. Edit users.json to add users.")
    st.stop()

username = st.session_state["username"]
role = st.session_state["role"]

# header
col1, col2 = st.columns([8, 1])
with col1:
    if os.path.exists("assets/logo.png"):
        st.image("assets/logo.png", width=120)
    st.markdown("## :triangular_ruler: BuildWise - Smart Estimation Tool")
with col2:
    if st.button("Logout"):
        do_logout()
        st.experimental_rerun()

# ensure folders
user_folder = os.path.join("user_data", username)
os.makedirs(user_folder, exist_ok=True)
os.makedirs(FORM_FOLDER, exist_ok=True)
os.makedirs(os.path.join(user_folder, "quotations"), exist_ok=True)

# ------------------------------
# Match settings (per-user)
# ------------------------------
def weights_file_for(user):
    folder = os.path.join("user_data", user)
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "weights.json")

def load_weights_for(user):
    path = weights_file_for(user)
    defaults = {"threshold": 70, "size": 0.45, "cores": 0.25, "material": 0.30}
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return {
                "threshold": int(data.get("threshold", defaults["threshold"])),
                "size": float(data.get("size", defaults["size"])),
                "cores": float(data.get("cores", defaults["cores"])),
                "material": float(data.get("material", defaults["material"])),
            }
        except Exception:
            return defaults
    return defaults

def save_weights_for(user, wdict):
    path = weights_file_for(user)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(wdict, f, indent=2, ensure_ascii=False)

if "match_settings_loaded" not in st.session_state:
    st.session_state["match_settings_loaded"] = True
    ws = load_weights_for(username)
    st.session_state["match_threshold"] = ws["threshold"]
    st.session_state["weight_size"] = ws["size"]
    st.session_state["weight_cores"] = ws["cores"]
    st.session_state["weight_material"] = ws["material"]

# ------------------------------
# Customers / trading terms
# ------------------------------
def user_customers_file(user):
    folder = os.path.join("user_data", user)
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "customers.json")

def load_customers_for(user):
    path = user_customers_file(user)
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_customers_for(user, customers):
    path = user_customers_file(user)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(customers, f, indent=2, ensure_ascii=False)

def trading_terms_file(user):
    folder = os.path.join("user_data", user)
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "trading_terms.json")

def load_trading_terms(user):
    path = trading_terms_file(user)
    defaults = {
        "payment": "",
        "delivery": "",
        "transportation_fee": "",
        "validity": "",
    }
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                d = json.load(f)
            for k in defaults:
                defaults[k] = d.get(k, defaults[k])
            return defaults
        except Exception:
            return defaults
    return defaults

def save_trading_terms(user, data):
    path = trading_terms_file(user)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

# ------------------------------
# Price list helpers
# ------------------------------
def list_price_list_files(folder_path):
    try:
        return sorted(
            f
            for f in os.listdir(folder_path)
            if os.path.isfile(os.path.join(folder_path, f))
            and f.lower().endswith((".xlsx", ".xls"))
            and f != TEMPLATE_FILE
        )
    except Exception:
        return []

# ------------------------------
# Quotation helpers using template
# ------------------------------
def generate_quotation_from_template(
    template_path,
    result_df,
    company_info,
    customer_info,
    trading_terms,
):
    """
    Fill quotation_template.xlsx:
    - Sheet1 (Quotation): company, customer, terms, date, quotation number, grand total (G12)
    - Sheet2 (Matched items): data from result_df starting at A2, sum Total -> grand total
    Returns: bytes of filled workbook.
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Quotation template '{template_path}' not found. Place it next to this app."
        )

    wb = load_workbook(template_path)
    # Try to get sheets by name, fallback by index
    try:
        ws_quote = wb["Quotation"]
    except KeyError:
        ws_quote = wb.worksheets[0]
    try:
        ws_items = wb["Matched items"]
    except KeyError:
        ws_items = wb.worksheets[1]

    # company info
    ws_quote["B2"] = company_info.get("name", "")
    ws_quote["B3"] = company_info.get("address", "")
    ws_quote["B4"] = company_info.get("phone", "")
    ws_quote["B5"] = company_info.get("email", "")
    ws_quote["B6"] = datetime.now().strftime("%Y-%m-%d")
    ws_quote["B7"] = "QT-" + datetime.now().strftime("%Y%m%d-%H%M%S")

    # customer info
    ws_quote["E3"] = customer_info.get("company", "")
    ws_quote["E4"] = customer_info.get("name", "")
    ws_quote["E5"] = customer_info.get("phone", "")
    ws_quote["E6"] = customer_info.get("email", "")
    ws_quote["E7"] = customer_info.get("address", "")

    # trading terms
    ws_quote["C21"] = trading_terms.get("payment", "")
    ws_quote["C22"] = trading_terms.get("delivery", "")
    ws_quote["C23"] = trading_terms.get("transportation_fee", "")
    ws_quote["C24"] = trading_terms.get("validity", "")

    # matched items in sheet2, starting A2
    # exclude last row if it is grand total row
    df = result_df.copy()
    if len(df) > 0:
        # detect if last row is grand total row
        last = df.iloc[-1]
        if (
            str(last.get("Model", "")).strip() == ""
            and str(last.get("Description (requested)", "")).strip() == ""
        ):
            df = df.iloc[:-1]

    start_row = 2
    # optional: clear existing rows (simple approach)
    max_rows_to_clear = 500
    for r in range(start_row, start_row + max_rows_to_clear):
        for c in range(1, 12):
            ws_items.cell(row=r, column=c, value=None)

    grand_total = 0.0
    for idx, (_, r) in enumerate(df.iterrows()):
        excel_row = start_row + idx
        ws_items.cell(row=excel_row, column=1, value=r.get("Model", ""))
        ws_items.cell(
            row=excel_row,
            column=2,
            value=r.get("Description (requested)", ""),
        )
        ws_items.cell(
            row=excel_row,
            column=3,
            value=r.get("Description (proposed)", ""),
        )
        ws_items.cell(row=excel_row, column=4, value=r.get("Specification", ""))
        ws_items.cell(row=excel_row, column=5, value=r.get("Unit", ""))
        ws_items.cell(row=excel_row, column=6, value=r.get("Quantity", 0))
        ws_items.cell(row=excel_row, column=7, value=r.get("Material Cost", 0))
        ws_items.cell(row=excel_row, column=8, value=r.get("Labour Cost", 0))
        ws_items.cell(
            row=excel_row, column=9, value=r.get("Amount Material", 0)
        )
        ws_items.cell(
            row=excel_row, column=10, value=r.get("Amount Labour", 0)
        )
        total_val = r.get("Total", 0)
        try:
            total_num = float(total_val)
        except Exception:
            total_num = 0.0
        ws_items.cell(row=excel_row, column=11, value=total_num)
        grand_total += total_num

    # write grand total to Sheet1 G12
    ws_quote["G12"] = grand_total

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def make_quotation_filename():
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"Quotation_{ts}.xlsx"

# ------------------------------
# Company profile
# ------------------------------
def page_company_profile():
    st.subheader("🏢 Company Profile")
    comp_file = os.path.join(user_folder, "company.json")
    profile = {}
    if os.path.exists(comp_file):
        try:
            with open(comp_file, "r", encoding="utf-8") as f:
                profile = json.load(f)
        except Exception:
            profile = {}

    name = st.text_input("Company name", value=profile.get("name", ""))
    address = st.text_input("Address", value=profile.get("address", ""))
    phone = st.text_input("Phone", value=profile.get("phone", ""))
    email = st.text_input("Email", value=profile.get("email", ""))

    if st.button("Save company profile"):
        data = {
            "name": name.strip(),
            "address": address.strip(),
            "phone": phone.strip(),
            "email": email.strip(),
        }
        with open(comp_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        st.success("Company profile saved.")

# ------------------------------
# Customers page (fixed edit/save)
# ------------------------------
def page_customers():
    st.subheader("👥 Customers")

    if role == "admin":
        st.info("Admin: view customers of each user.")
        base = "user_data"
        os.makedirs(base, exist_ok=True)
        user_dirs = sorted(
            d for d in os.listdir(base) if os.path.isdir(os.path.join(base, d))
        )
        chosen_user = st.selectbox("Select user", ["--Select--"] + user_dirs)
        if chosen_user == "--Select--":
            return
        owner = chosen_user
    else:
        owner = username

    customers = load_customers_for(owner)

    with st.expander("Add new customer", expanded=False):
        with st.form("add_customer_form"):
            c_name = st.text_input("Customer name")
            c_company = st.text_input("Company")
            c_address = st.text_input("Address")
            c_phone = st.text_input("Phone")
            c_email = st.text_input("Email")
            c_notes = st.text_area("Notes")
            submitted = st.form_submit_button("Add customer")
            if submitted:
                if not c_name.strip():
                    st.error("Customer name is required.")
                else:
                    new = {
                        "id": f"C{int(datetime.now().timestamp())}",
                        "name": c_name.strip(),
                        "company": c_company.strip(),
                        "address": c_address.strip(),
                        "phone": c_phone.strip(),
                        "email": c_email.strip(),
                        "notes": c_notes.strip(),
                        "created_at": datetime.now().isoformat(),
                    }
                    customers.append(new)
                    save_customers_for(owner, customers)
                    st.success("Customer added.")
                    st.experimental_rerun()

    if not customers:
        st.info("No customers yet.")
        return

    df = pd.DataFrame(customers)
    cols_order = [
        "id",
        "name",
        "company",
        "phone",
        "email",
        "address",
        "notes",
        "created_at",
    ]
    cols = [c for c in cols_order if c in df.columns] + [
        c for c in df.columns if c not in cols_order
    ]
    df = df[cols]
    st.markdown("### Customer list")
    st.dataframe(df.reset_index(drop=True), use_container_width=True)

    st.markdown("### Edit or delete customer")
    ids = df["id"].astype(str).tolist()
    sel_id = st.selectbox("Select customer ID", [""] + ids)
    if not sel_id:
        return

    selected = df[df["id"].astype(str) == sel_id].iloc[0].to_dict()

    col_edit, col_delete = st.columns([3, 1])

    with col_edit:
        with st.form(f"edit_customer_form_{sel_id}"):
            e_name = st.text_input("Customer name", value=selected.get("name", ""))
            e_company = st.text_input("Company", value=selected.get("company", ""))
            e_address = st.text_input("Address", value=selected.get("address", ""))
            e_phone = st.text_input("Phone", value=selected.get("phone", ""))
            e_email = st.text_input("Email", value=selected.get("email", ""))
            e_notes = st.text_area("Notes", value=selected.get("notes", ""))
            submitted = st.form_submit_button("Save customer")
            if submitted:
                for i, c in enumerate(customers):
                    if str(c.get("id")) == sel_id:
                        customers[i].update(
                            {
                                "name": e_name.strip(),
                                "company": e_company.strip(),
                                "address": e_address.strip(),
                                "phone": e_phone.strip(),
                                "email": e_email.strip(),
                                "notes": e_notes.strip(),
                                "updated_at": datetime.now().isoformat(),
                            }
                        )
                        break
                save_customers_for(owner, customers)
                st.success("Customer updated.")
                st.experimental_rerun()

    with col_delete:
        if st.button("Delete customer"):
            new_list = [c for c in customers if str(c.get("id")) != sel_id]
            save_customers_for(owner, new_list)
            st.success("Customer deleted.")
            st.experimental_rerun()

# ------------------------------
# Forms & Instructions
# ------------------------------
def page_forms_and_instructions():
    st.subheader("📂 Forms and Instructions")
    st.write("Shared templates and instructions.")

    form_files = sorted(os.listdir(FORM_FOLDER))

    if role == "admin":
        uploads = st.file_uploader(
            "Admin: Upload forms (xlsx/xls)",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
        )
        if uploads:
            for f in uploads:
                path = os.path.join(FORM_FOLDER, f.name)
                try:
                    with open(path, "wb") as out_f:
                        out_f.write(f.read())
                except Exception as e:
                    st.error(f"Error saving {f.name}: {e}")
            st.success("Form(s) uploaded.")
            st.experimental_rerun()

        if form_files:
            to_del = st.selectbox("Select form to delete", [""] + form_files)
            if to_del and st.button("Delete selected form"):
                try:
                    os.remove(os.path.join(FORM_FOLDER, to_del))
                    st.success("Form deleted.")
                    st.experimental_rerun()
                except Exception as e:
                    st.error(f"Error deleting form: {e}")
    else:
        if form_files:
            for f in form_files:
                path = os.path.join(FORM_FOLDER, f)
                try:
                    with open(path, "rb") as fh:
                        data = fh.read()
                    st.download_button(
                        f"Download {f}",
                        data,
                        file_name=f,
                        key=f"down_form_{f}",
                    )
                except Exception:
                    continue
        else:
            st.info("No forms available.")

# ------------------------------
# Quotations page
# ------------------------------
def page_quotations():
    st.subheader("📄 Quotations")
    q_folder = os.path.join(user_folder, "quotations")
    os.makedirs(q_folder, exist_ok=True)
    files = sorted(os.listdir(q_folder))
    if not files:
        st.info("No quotations saved.")
        return

    for f in files:
        path = os.path.join(q_folder, f)
        c1, c2, c3 = st.columns([4, 1, 1])
        with c1:
            st.write(f)
        with c2:
            with open(path, "rb") as fh:
                data = fh.read()
            st.download_button(
                "Download", data, file_name=f, key=f"down_q_{f}"
            )
        with c3:
            if st.button("Delete", key=f"del_q_{f}"):
                os.remove(path)
                st.success("Quotation deleted.")
                st.experimental_rerun()

# ------------------------------
# Estimation page
# ------------------------------
def page_estimation():
    # session state for matching & quotation
    if "last_match_df" not in st.session_state:
        st.session_state["last_match_df"] = None
    if "last_unmatched_df" not in st.session_state:
        st.session_state["last_unmatched_df"] = None
    if "quotation_bytes" not in st.session_state:
        st.session_state["quotation_bytes"] = None
    if "quotation_filename" not in st.session_state:
        st.session_state["quotation_filename"] = None

    st.subheader("1. Upload price list files")
    uploads = st.file_uploader(
        "Upload one or more price list Excel files (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=True,
        key="pl_up_main",
    )
    if uploads:
        for f in uploads:
            try:
                with open(os.path.join(user_folder, f.name), "wb") as out_f:
                    out_f.write(f.read())
            except Exception as e:
                st.error(f"Error saving {f.name}: {e}")
        st.success("Price list(s) uploaded.")
        st.experimental_rerun()

    st.subheader("2. Manage price lists")
    price_list_files = list_price_list_files(user_folder)
    if price_list_files:
        st.write("Your price lists:")
        for f in price_list_files:
            st.write(f"- {f}")
    else:
        st.info("No price lists uploaded yet.")

    selected_file = st.radio(
        "Choose one price list to use or all",
        ["All files"] + price_list_files,
        index=0,
    )

    if price_list_files:
        col_a, col_b = st.columns([3, 1])
        with col_a:
            to_del = st.selectbox(
                "Select a price list to delete",
                [""] + price_list_files,
                key="del_pl_main",
            )
        with col_b:
            if st.button("Delete selected price list"):
                if to_del:
                    try:
                        os.remove(os.path.join(user_folder, to_del))
                        st.success(f"Deleted {to_del}")
                        st.experimental_rerun()
                    except Exception as e:
                        st.error(f"Error deleting file: {e}")

    st.markdown("---")
    st.subheader("3. Matching estimation request file")

    estimation_file = st.file_uploader(
        "Upload estimation request (.xlsx)",
        type=["xlsx"],
        key="estimation_file_main",
    )

    match_threshold = st.session_state.get("match_threshold", 70)
    w_size = st.session_state.get("weight_size", 0.45)
    w_cores = st.session_state.get("weight_cores", 0.25)
    w_material = st.session_state.get("weight_material", 0.30)
    total_w = w_size + w_cores + w_material
    if total_w <= 0:
        total_w = 1.0
    weights = {
        "size": w_size / total_w,
        "cores": w_cores / total_w,
        "material": w_material / total_w,
    }

    col_match_btn, _ = st.columns([1, 3])
    with col_match_btn:
        run_matching = st.button("Match now")

    if run_matching:
        if estimation_file is None:
            st.error("Please upload an estimation file first.")
        elif not price_list_files:
            st.error("Please upload at least one price list first.")
        else:
            # read estimation
            try:
                est = pd.read_excel(estimation_file).dropna(how="all")
            except Exception as e:
                st.error(f"Cannot read estimation file: {e}")
                est = None

            if est is not None:
                est_cols = est.columns.tolist()
                if len(est_cols) < 5:
                    st.error(
                        "Estimation file must have at least 5 columns (Model, Description, Spec, Unit, Quantity)."
                    )
                else:
                    base_est = (
                        est[est_cols[0]].fillna("")
                        + " "
                        + est[est_cols[1]].fillna("")
                        + " "
                        + est[est_cols[2]].fillna("")
                    )
                    est["combined"] = base_est.apply(clean)
                    parsed_est = base_est.apply(parse_cable_spec)
                    est["main_key"] = parsed_est.apply(lambda d: d["main_key"])
                    est["aux_key"] = parsed_est.apply(lambda d: d["aux_key"])
                    est["materials"] = base_est.apply(
                        extract_material_structure_tokens
                    )

                    # read DB(s)
                    if selected_file == "All files":
                        frames = []
                        for f in price_list_files:
                            try:
                                df_pl = pd.read_excel(
                                    os.path.join(user_folder, f)
                                ).dropna(how="all")
                                df_pl["source"] = f
                                frames.append(df_pl)
                            except Exception:
                                continue
                        db = (
                            pd.concat(frames, ignore_index=True)
                            if frames
                            else pd.DataFrame()
                        )
                    else:
                        try:
                            db = pd.read_excel(
                                os.path.join(user_folder, selected_file)
                            ).dropna(how="all")
                            db["source"] = selected_file
                        except Exception as e:
                            st.error(f"Cannot read price list: {e}")
                            db = pd.DataFrame()

                    if db.empty:
                        st.error("No rows found in price list file(s).")
                    else:
                        db_cols = db.columns.tolist()
                        if len(db_cols) < 6:
                            st.error(
                                "Price list requires at least 6 columns (Model, Description, Spec, ..., MaterialCost, LabourCost)."
                            )
                        else:
                            base_db = (
                                db[db_cols[0]].fillna("")
                                + " "
                                + db[db_cols[1]].fillna("")
                                + " "
                                + db[db_cols[2]].fillna("")
                            )
                            db["combined"] = base_db.apply(clean)
                            parsed_db = base_db.apply(parse_cable_spec)
                            db["main_key"] = parsed_db.apply(
                                lambda d: d["main_key"]
                            )
                            db["aux_key"] = parsed_db.apply(
                                lambda d: d["aux_key"]
                            )
                            db["materials"] = base_db.apply(
                                extract_material_structure_tokens
                            )

                            results = []
                            for _, row in est.iterrows():
                                query = row["combined"]
                                q_main = row["main_key"]
                                q_aux = row["aux_key"]
                                q_mats = row["materials"]
                                unit = row[est_cols[3]]
                                qty_value = row[est_cols[4]]

                                best = None
                                best_score = -1.0

                                c0 = db.copy()
                                if q_main:
                                    c0 = c0[c0["main_key"] == q_main]

                                def score_row(r):
                                    try:
                                        r_main = r.get("main_key", "")
                                        r_aux = r.get("aux_key", "")
                                        r_mats = r.get("materials", [])
                                        return combined_match_score(
                                            query,
                                            q_main,
                                            q_aux,
                                            q_mats,
                                            r.get("combined", ""),
                                            r_main,
                                            r_aux,
                                            r_mats,
                                            match_threshold,
                                            weights,
                                        )
                                    except Exception:
                                        return 0.0

                                if not c0.empty:
                                    c0 = c0.copy()
                                    c0["score"] = c0.apply(score_row, axis=1)
                                    top = (
                                        c0.sort_values(
                                            "score", ascending=False
                                        )
                                        .head(1)
                                        .reset_index(drop=True)
                                    )
                                    if (
                                        not top.empty
                                        and float(top.loc[0, "score"])
                                        >= match_threshold
                                    ):
                                        best = top.loc[0]
                                        best_score = float(best["score"])

                                if best is None:
                                    c1 = db.copy()
                                    c1["score"] = c1.apply(score_row, axis=1)
                                    top2 = (
                                        c1.sort_values(
                                            "score", ascending=False
                                        )
                                        .head(1)
                                        .reset_index(drop=True)
                                    )
                                    if (
                                        not top2.empty
                                        and float(top2.loc[0, "score"])
                                        >= match_threshold
                                    ):
                                        best = top2.loc[0]
                                        best_score = float(best["score"])

                                if best is None:
                                    c2 = db.copy()
                                    c2["score"] = c2["combined"].apply(
                                        lambda x: fuzz.token_set_ratio(
                                            query, x
                                        )
                                    )
                                    top3 = (
                                        c2.sort_values(
                                            "score", ascending=False
                                        )
                                        .head(1)
                                        .reset_index(drop=True)
                                    )
                                    if not top3.empty:
                                        best = top3.loc[0]
                                        best_score = float(best["score"])

                                if best is not None and best_score >= 0:
                                    matched_desc = best[db_cols[1]]
                                    matched_model = best[db_cols[0]]
                                    matched_spec = best[db_cols[2]]
                                    m_cost = pd.to_numeric(
                                        best[db_cols[4]], errors="coerce"
                                    )
                                    l_cost = pd.to_numeric(
                                        best[db_cols[5]], errors="coerce"
                                    )
                                    if pd.isna(m_cost):
                                        m_cost = 0
                                    if pd.isna(l_cost):
                                        l_cost = 0
                                else:
                                    matched_desc = ""
                                    matched_model = ""
                                    matched_spec = ""
                                    m_cost = 0
                                    l_cost = 0

                                qty_num = pd.to_numeric(
                                    qty_value, errors="coerce"
                                )
                                if pd.isna(qty_num):
                                    qty_num = 0
                                amt_mat = qty_num * m_cost
                                amt_lab = qty_num * l_cost
                                total_val = amt_mat + amt_lab

                                results.append(
                                    [
                                        matched_model,
                                        row[est_cols[1]],
                                        matched_desc,
                                        matched_spec,
                                        unit,
                                        qty_num,
                                        m_cost,
                                        l_cost,
                                        amt_mat,
                                        amt_lab,
                                        total_val,
                                    ]
                                )

                            result_df = pd.DataFrame(
                                results,
                                columns=[
                                    "Model",
                                    "Description (requested)",
                                    "Description (proposed)",
                                    "Specification",
                                    "Unit",
                                    "Quantity",
                                    "Material Cost",
                                    "Labour Cost",
                                    "Amount Material",
                                    "Amount Labour",
                                    "Total",
                                ],
                            )
                            grand_total = pd.to_numeric(
                                result_df["Total"], errors="coerce"
                            ).sum()
                            result_df.loc[len(result_df.index)] = (
                                [""] * 10 + [grand_total]
                            )

                            unmatched_df = result_df[
                                result_df["Description (proposed)"] == ""
                            ]

                            st.session_state["last_match_df"] = result_df
                            st.session_state["last_unmatched_df"] = unmatched_df
                            st.success("Matching completed.")

    # show last matching results (persistent while editing)
    last_df = st.session_state.get("last_match_df")
    last_unmatched = st.session_state.get("last_unmatched_df")

    if last_df is not None:
        st.markdown("#### Matched Estimation (latest)")
        display_df = last_df.copy()
        display_df["Quantity"] = pd.to_numeric(
            display_df["Quantity"], errors="coerce"
        ).fillna(0).astype(int)
        for col in [
            "Material Cost",
            "Labour Cost",
            "Amount Material",
            "Amount Labour",
            "Total",
        ]:
            display_df[col] = pd.to_numeric(
                display_df[col], errors="coerce"
            ).fillna(0).map("{:,.0f}".format)
        st.dataframe(display_df, use_container_width=True)

        # download matching file (.xlsx)
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            last_df.to_excel(
                writer, index=False, sheet_name="Matched Results"
            )
            if last_unmatched is not None and not last_unmatched.empty:
                last_unmatched.to_excel(
                    writer,
                    index=False,
                    sheet_name="Unmatched Items",
                )
        st.download_button(
            "Download matching file (.xlsx)",
            buffer.getvalue(),
            file_name="Estimation_Result_BuildWise.xlsx",
        )
    else:
        st.info("No matching result yet. Upload files and click 'Match now'.")

    st.markdown("---")
    st.subheader("4. Quotation generation")

    # customers
    customers = load_customers_for(username)
    cust_labels = ["--No customer--"] + [
        f"{c.get('name', '')} ({c.get('company', '')})" for c in customers
    ]
    col_c1, col_c2 = st.columns([2, 2])
    with col_c1:
        selected_cust_label = st.selectbox(
            "Select a customer", cust_labels, index=0
        )

    active_customer = None
    if selected_cust_label != "--No customer--":
        idx = cust_labels.index(selected_cust_label) - 1
        active_customer = customers[idx]
        st.markdown("*Selected customer:*")
        # display as table
        cust_df = pd.DataFrame(
            {
                "Field": [
                    "Name",
                    "Company",
                    "Address",
                    "Phone",
                    "Email",
                    "Notes",
                ],
                "Value": [
                    active_customer.get("name", ""),
                    active_customer.get("company", ""),
                    active_customer.get("address", ""),
                    active_customer.get("phone", ""),
                    active_customer.get("email", ""),
                    active_customer.get("notes", ""),
                ],
            }
        )
        st.table(cust_df)
    else:
        st.info("No customer selected yet.")

    # Edit/save selected customer (fixed)
    if active_customer is not None:
        with st.expander("Edit selected customer"):
            with st.form("edit_selected_customer_main"):
                e_name = st.text_input(
                    "Customer name", value=active_customer.get("name", "")
                )
                e_company = st.text_input(
                    "Company", value=active_customer.get("company", "")
                )
                e_address = st.text_input(
                    "Address", value=active_customer.get("address", "")
                )
                e_phone = st.text_input(
                    "Phone", value=active_customer.get("phone", "")
                )
                e_email = st.text_input(
                    "Email", value=active_customer.get("email", "")
                )
                e_notes = st.text_area(
                    "Notes", value=active_customer.get("notes", "")
                )
                submitted = st.form_submit_button("Save customer")
                if submitted:
                    cust_id = active_customer.get("id")
                    for i, c in enumerate(customers):
                        if c.get("id") == cust_id:
                            customers[i].update(
                                {
                                    "name": e_name.strip(),
                                    "company": e_company.strip(),
                                    "address": e_address.strip(),
                                    "phone": e_phone.strip(),
                                    "email": e_email.strip(),
                                    "notes": e_notes.strip(),
                                    "updated_at": datetime.now().isoformat(),
                                }
                            )
                            break
                    save_customers_for(username, customers)
                    st.success("Customer updated.")
                    st.experimental_rerun()

    # Trading terms
    st.markdown("#### Trading terms / Điều khoản thương mại")
    terms = load_trading_terms(username)
    with st.form("trading_terms_form_main"):
        payment = st.text_area(
            "Payment / Thanh toán",
            value=terms.get("payment", ""),
            height=80,
        )
        delivery = st.text_input(
            "Delivery schedule / Tiến độ",
            value=terms.get("delivery", ""),
        )
        trans_fee = st.text_input(
            "Transportation fee / Phí vận chuyển",
            value=terms.get("transportation_fee", ""),
        )
        validity = st.text_input(
            "Quotation validity / Hiệu lực báo giá",
            value=terms.get("validity", ""),
        )
        save_terms_btn = st.form_submit_button("Save trading terms")
        if save_terms_btn:
            new_terms = {
                "payment": payment,
                "delivery": delivery,
                "transportation_fee": trans_fee,
                "validity": validity,
            }
            save_trading_terms(username, new_terms)
            st.success("Trading terms saved.")
            terms = new_terms

    col_g1, col_g2 = st.columns([1, 3])
    with col_g1:
        generate_q = st.button("Generate quotation")

    if generate_q:
        if active_customer is None:
            st.error("Please select a customer before generating quotation.")
        elif st.session_state.get("last_match_df") is None:
            st.error("Please run matching first.")
        elif not os.path.exists(TEMPLATE_FILE):
            st.error(
                f"Quotation template '{TEMPLATE_FILE}' not found. Please upload it to the same folder as this app."
            )
        else:
            # load company info
            comp_file = os.path.join(user_folder, "company.json")
            company_info = {}
            if os.path.exists(comp_file):
                try:
                    with open(comp_file, "r", encoding="utf-8") as f:
                        company_info = json.load(f)
                except Exception:
                    company_info = {}

            result_df = st.session_state["last_match_df"].copy()
            current_terms = {
                "payment": payment,
                "delivery": delivery,
                "transportation_fee": trans_fee,
                "validity": validity,
            }
            save_trading_terms(username, current_terms)

            try:
                q_bytes = generate_quotation_from_template(
                    TEMPLATE_FILE,
                    result_df,
                    company_info,
                    active_customer,
                    current_terms,
                )
                q_filename = make_quotation_filename()
                st.session_state["quotation_bytes"] = q_bytes
                st.session_state["quotation_filename"] = q_filename
                st.success("Quotation generated using template.")
            except FileNotFoundError as e:
                st.error(str(e))
            except Exception as e:
                st.error(f"Error generating quotation: {e}")

    # Quotation preview + Download + Save
    if st.session_state.get("quotation_bytes") is not None:
        st.markdown("#### Quotation preview (matched items)")
        # preview from last_match_df
        prev_df = st.session_state["last_match_df"].copy()
        display_prev = prev_df.copy()
        display_prev["Quantity"] = pd.to_numeric(
            display_prev["Quantity"], errors="coerce"
        ).fillna(0).astype(int)
        for col in [
            "Material Cost",
            "Labour Cost",
            "Amount Material",
            "Amount Labour",
            "Total",
        ]:
            display_prev[col] = pd.to_numeric(
                display_prev[col], errors="coerce"
            ).fillna(0).map("{:,.0f}".format)
        st.dataframe(display_prev, use_container_width=True)

        col_d1, col_d2 = st.columns([1, 1])
        with col_d1:
            st.download_button(
                "Download quotation (.xlsx)",
                st.session_state["quotation_bytes"],
                file_name=st.session_state["quotation_filename"],
            )
        with col_d2:
            if st.button("Save quotation"):
                q_folder = os.path.join(user_folder, "quotations")
                os.makedirs(q_folder, exist_ok=True)
                path = os.path.join(
                    q_folder, st.session_state["quotation_filename"]
                )
                with open(path, "wb") as f:
                    f.write(st.session_state["quotation_bytes"])
                st.success("Quotation saved to history.")
    else:
        st.info("Generate a quotation to enable preview, download, and save.")

# ------------------------------
# Sidebar navigation + match settings
# ------------------------------
st.sidebar.title("Navigation")
nav_items = [
    "Estimation",
    "Customers",
    "Company Profile",
    "Quotation",
    "Forms and Instructions",
]
page = st.sidebar.radio("Go to", nav_items, index=0)

st.sidebar.markdown("---")
st.sidebar.subheader("Matching settings")
th = st.sidebar.slider(
    "Match threshold", 0, 100, st.session_state.get("match_threshold", 70)
)
w_s = st.sidebar.slider(
    "Size weight",
    0.0,
    1.0,
    st.session_state.get("weight_size", 0.45),
    step=0.05,
)
w_c = st.sidebar.slider(
    "Cores weight",
    0.0,
    1.0,
    st.session_state.get("weight_cores", 0.25),
    step=0.05,
)
w_m = st.sidebar.slider(
    "Material weight",
    0.0,
    1.0,
    st.session_state.get("weight_material", 0.30),
    step=0.05,
)

if st.sidebar.button("Save matching settings"):
    settings = {
        "threshold": int(th),
        "size": float(w_s),
        "cores": float(w_c),
        "material": float(w_m),
    }
    save_weights_for(username, settings)
    st.session_state["match_threshold"] = settings["threshold"]
    st.session_state["weight_size"] = settings["size"]
    st.session_state["weight_cores"] = settings["cores"]
    st.session_state["weight_material"] = settings["material"]
    st.sidebar.success("Matching settings saved.")

# ------------------------------
# Routing
# ------------------------------
if page == "Estimation":
    page_estimation()
elif page == "Customers":
    page_customers()
elif page == "Company Profile":
    page_company_profile()
elif page == "Quotation":
    page_quotations()
elif page == "Forms and Instructions":
    page_forms_and_instructions()

st.markdown("---")
st.caption("BuildWise — Estimation & Quotation tool (Excel template version)")
